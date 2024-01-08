import openpyxl
from utils.constants import (
    EXCEL_FILE,
    TASKS_SHEET,
    HISTORY_SHEET,
    DESCRIPTION,
    DATE,
    START_TIME,
    END_TIME,
    DEDICATED_TIME,
    TASK_TW,
)
from datetime import datetime


def obtain_tasks_info():
    print("Obtaining tasks info...")
    workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet_tasks = workbook[TASKS_SHEET]

    tasks_info = []

    for row in sheet_tasks.iter_rows(min_row=2): # Se asume que la primera fila tiene los encabezados
        try:
            dedicated_time = row[DEDICATED_TIME - 1].value.strftime("%H:%M") # Se resta 1 porque openpyxl empieza a contar desde 1
        except Exception as e:
            if str(e) == "'NoneType' object has no attribute 'strftime'":
                continue
            else:
                print(e)

        description = row[DESCRIPTION - 1].value
        date = datetime.strptime(str(row[DATE - 1].value), "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
        start_time = row[START_TIME - 1].value.strftime("%I:%M %p")
        end_time = row[END_TIME - 1].value.strftime("%I:%M %p")
        task_tw = row[TASK_TW - 1].value
        task = {
            "date": str(date),
            "start_time": start_time,
            "end_time": end_time,
            "dedicated_time": dedicated_time,
            "task_tw": task_tw,
            "description": description,
            "index": row[0].row, # Se obtiene el número de fila de la primera celda
        }
        tasks_info.append(task)
        print(task.values())

    print("Tasks info obtained")
    return tasks_info


def tasks_process(tasks_info):
    print("Starting tasks process...")
    finished_tasks = []
    for task in tasks_info:
        print(task)
        if task["task_tw"] == "#N/A":
            print("Task not found")
            continue
        finished_tasks.append(task)
    print("Tasks process finished")
    return finished_tasks


def move_finished_tasks_to_history(finished_tasks):
    print("Moving finished tasks to history...")
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet_tasks = workbook[TASKS_SHEET]
    sheet_history = workbook[HISTORY_SHEET]

    for task in finished_tasks:
        if task["index"] > 28:
            row_values = [cell.value for cell in sheet_tasks[task["index"]]] # Se obtiene una lista con los valores de la fila
            sheet_history.append(row_values) # Se agrega la fila al final de la hoja de historial
            # sheet_tasks.delete_rows(task["index"]) # Se elimina la fila de la hoja de tareas

    workbook.save(EXCEL_FILE) # Se guarda el archivo de Excel
    print("Finished tasks moved to history")


def start_process():
    print("Starting process...")
    tasks_info = obtain_tasks_info()
    tasks_info.reverse()

    finised_tasks = tasks_process(tasks_info)
    print("Process finished")
    print(finised_tasks)
    move_finished_tasks_to_history(finised_tasks)


start_process()
